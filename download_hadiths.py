#!/usr/bin/env python3
"""
download_hadiths.py
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
Descarga hadices en espaÃ±ol desde HadeethEnc.com y genera archivos JSON
para integrar en la app Flutter Qibla Time.

USO:
    pip install requests openpyxl
    python download_hadiths.py

ARCHIVOS GENERADOS:
    assets/hadiths/bukhari.json
    assets/hadiths/muslim.json
    assets/hadiths/tirmidhi.json
    etc.

FORMATO DE SALIDA:
    [
      {
        "id": 1,
        "arabic": "texto en Ã¡rabe",
        "translation": "traducciÃ³n al espaÃ±ol con tildes",
        "reference": "Bukhari 1",
        "category": "categoria",
        "grade": "Sahih"
      }
    ]
"""

import json
import os
import re
import time
import requests
import zipfile
import openpyxl
from pathlib import Path
from io import BytesIO

# â”€â”€ ConfiguraciÃ³n â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# URL de descarga de HadeethEnc (Excel en espaÃ±ol)
HADEETHENC_DOWNLOAD_URL = "https://hadeethenc.com/en/browse/download/es"

# Directorio de salida
OUTPUT_DIR = Path("assets/hadiths")

# Mapeo de nombres de colecciones (inglÃ©s â†’ nombre para archivo)
COLLECTION_MAPPING = {
    "Sahih al-Bukhari": "bukhari",
    "Sahih Muslim": "muslim",
    "Sunan Abi Dawud": "abudawud",
    "Jami` at-Tirmidhi": "tirmidhi",
    "Sunan an-Nasa'i": "nasai",
    "Sunan Ibn Majah": "ibnmajah",
    "Muwatta Malik": "malik",
    "Musnad Ahmad": "ahmad",
}

# â”€â”€ Funciones â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def download_excel() -> bytes | None:
    """Descarga el archivo Excel de HadeethEnc."""
    print("Descargando hadices desde HadeethEnc.com...")
    print(f"URL: {HADEETHENC_DOWNLOAD_URL}")

    try:
        response = requests.get(HADEETHENC_DOWNLOAD_URL, timeout=60)
        response.raise_for_status()

        # Verificar que es un archivo ZIP/Excel
        if not response.content.startswith(b'PK'):
            print("âŒ Error: El archivo descargado no parece ser un ZIP/Excel")
            return None

        print(f"âœ… Descargado: {len(response.content) / (1024*1024):.1f} MB")
        return response.content

    except Exception as e:
        print(f"âŒ Error descargando: {e}")
        return None


def extract_excel(zip_data: bytes) -> BytesIO | None:
    """Extrae el archivo Excel del ZIP (puede ser ZIP dentro de ZIP o Excel directo)."""
    try:
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zf:
            file_list = zf.namelist()
            print(f"Archivos en el ZIP: {len(file_list)} archivos")

            # Caso 1: Buscar hojas de cÃ¡lculo (Excel descomprimido)
            sheet_files = [f for f in file_list if 'worksheets/' in f and f.endswith('.xml')]
            if sheet_files:
                print("âœ… El ZIP contiene un Excel descomprimido (formato Office Open XML)")
                # El ZIP ES el Excel, lo pasamos directamente
                return BytesIO(zip_data)

            # Caso 2: Excel directo (.xlsx)
            excel_files = [f for f in file_list if f.endswith('.xlsx')]
            if excel_files:
                excel_filename = excel_files[0]
                print(f"Extrayendo: {excel_filename}")
                return BytesIO(zf.read(excel_filename))

            # Caso 3: ZIP dentro de ZIP (HadeethEnc usa esto a veces)
            zip_files = [f for f in file_list if f.endswith('.zip')]
            if zip_files:
                inner_zip_name = zip_files[0]
                print(f"Extrayendo ZIP interno: {inner_zip_name}")
                inner_zip_data = zf.read(inner_zip_name)

                # Extraer el ZIP interno recursivamente
                return extract_excel(inner_zip_data)

            print("âŒ No se encontrÃ³ archivo Excel en el ZIP")
            print(f"   Archivos encontrados: {file_list[:15]}")
            return None

    except Exception as e:
        print(f"âŒ Error extrayendo ZIP: {e}")
        import traceback
        traceback.print_exc()
        return None


def parse_excel(excel_data: BytesIO) -> list[dict]:
    """Parsea el Excel y extrae los hadices."""
    print("Procesando archivo Excel...")

    wb = openpyxl.load_workbook(excel_data, data_only=True)
    ws = wb.active

    hadiths = []

    # Estructura de columnas de HadeethEnc:
    # A: id
    # B: title_ar (tÃ­tulo en Ã¡rabe)
    # C: title (tÃ­tulo en espaÃ±ol)
    # D: hadith_text_ar (texto Ã¡rabe)
    # E: hadith_text (texto en espaÃ±ol)
    # F: explanation_ar
    # G: explanation
    # H: benefits_ar
    # I: benefits
    # J: grade_ar
    # K: takhrij_ar
    # L: grade (grado de autenticidad)
    # M: takhrij (fuente/colecciÃ³n)
    # N: lang
    # O: link

    for row_idx in range(2, ws.max_row + 1):  # Saltar fila 1 (headers)
        try:
            # Leer celdas
            hadith_id_raw = ws.cell(row=row_idx, column=1).value

            # Saltar fila de headers si se cuela
            if hadith_id_raw == 'id':
                continue

            hadith_id = hadith_id_raw
            title_es = ws.cell(row=row_idx, column=3).value
            arabic = ws.cell(row=row_idx, column=4).value
            translation = ws.cell(row=row_idx, column=5).value
            grade = ws.cell(row=row_idx, column=12).value
            takhrij = ws.cell(row=row_idx, column=13).value

            # Saltar filas vacÃ­as
            if not arabic or not translation:
                continue

            # Limpiar textos
            arabic = str(arabic).strip() if arabic else ""
            translation = str(translation).strip() if translation else ""
            title_es = str(title_es).strip() if title_es else ""

            # Extraer colecciÃ³n del takhrij
            collection = extract_collection_from_takhrij(takhrij)
            category = extract_category_from_title(title_es, arabic)

            # Normalizar grado de autenticidad
            grade_normalized = normalize_grade(str(grade) if grade else "")

            # Generar referencia (limpiar corchetes)
            reference = str(takhrij).strip('[] ') if takhrij else collection

            hadiths.append({
                "id": int(hadith_id) if hadith_id else row_idx,
                "arabic": arabic,
                "translation": translation,
                "reference": reference,
                "category": category,
                "grade": grade_normalized,
                "_collection": collection,  # Temporal para agrupar
            })

            if row_idx % 500 == 0:
                print(f"  Procesados {row_idx} hadices...", end="\r")

        except Exception as e:
            print(f"âš ï¸  Error en fila {row_idx}: {e}")
            continue

    print(f"\nâœ… Total procesados: {len(hadiths)} hadices")
    return hadiths


def extract_collection_from_takhrij(takhrij) -> str:
    """Extrae el nombre de la colecciÃ³n del campo takhrij."""
    if not takhrij:
        return "general"

    takhrij_str = str(takhrij).lower()

    # Mapeo de colecciones
    if "al-bujari" in takhrij_str or "al-bukhari" in takhrij_str or "bujari" in takhrij_str:
        return "bukhari"
    elif "muslim" in takhrij_str:
        return "muslim"
    elif "tirmidhi" in takhrij_str:
        return "tirmidhi"
    elif "abu-dawud" in takhrij_str or "abudawud" in takhrij_str or "abu dawud" in takhrij_str:
        return "abudawud"
    elif "nasai" in takhrij_str or "an-nasa'i" in takhrij_str:
        return "nasai"
    elif "ibn-majah" in takhrij_str or "ibn majah" in takhrij_str or "ibnmajah" in takhrij_str:
        return "ibnmajah"
    elif "malik" in takhrij_str or "muwatta" in takhrij_str:
        return "malik"
    elif "ahmad" in takhrij_str or "musnad ahmad" in takhrij_str:
        return "ahmad"
    elif "registrado por" in takhrij_str:
        # Si hay mÃºltiples, tomar la primera
        if "bujari" in takhrij_str:
            return "bukhari"
        elif "muslim" in takhrij_str:
            return "muslim"

    return "general"


def extract_category_from_title(title_es, arabic) -> str:
    """Extrae una categorÃ­a corta del tÃ­tulo del hadiz."""
    if not title_es:
        return "general"

    # Limpiar y acortar
    category = str(title_es).lower()
    # Quitar saltos de lÃ­nea
    category = category.replace('\n', ' ').replace('\r', ' ')
    # Acortar a 80 caracteres
    if len(category) > 80:
        category = category[:77] + "..."

    return category.strip()


def normalize_grade(grade: str) -> str:
    """Normaliza el grado de autenticidad al formato de la app."""
    if not grade:
        return "Desconocido"

    grade_lower = grade.lower().strip()

    # Mapeo de grados
    if any(x in grade_lower for x in ["sahih", "autÃ©ntico", "autentico", "correcto"]):
        return "Sahih"
    elif any(x in grade_lower for x in ["hasan", "bueno", "aceptable"]):
        return "Hasan"
    elif any(x in grade_lower for x in ["da'if", "daif", "dÃ©bil", "debil", "flojo"]):
        return "Da'if"
    elif any(x in grade_lower for x in ["mawdu'", "mawdu", "falso", "inventado"]):
        return "Mawdu'"
    elif any(x in grade_lower for x in ["muttafaq", "acordado", "bukhari muslim"]):
        return "Muttafaqun Alayhi"
    else:
        return grade.strip().title()


def group_by_collection(hadiths: list[dict]) -> dict[str, list[dict]]:
    """Agrupa los hadices por colecciÃ³n."""
    collections = {}

    for hadith in hadiths:
        collection_name = hadith.pop("_collection")  # Remover campo temporal

        # Normalizar nombre de colecciÃ³n
        collection_key = None
        for en_name, file_key in COLLECTION_MAPPING.items():
            if en_name.lower() in collection_name.lower():
                collection_key = file_key
                break

        if not collection_key:
            # Crear key genÃ©rico si no estÃ¡ en el mapeo
            collection_key = re.sub(r'[^a-z0-9]', '_', collection_name.lower()).strip('_')

        if collection_key not in collections:
            collections[collection_key] = []

        collections[collection_key].append(hadith)

    return collections


def save_collections(collections: dict[str, list[dict]]) -> None:
    """Guarda cada colecciÃ³n en un archivo JSON separado."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Guardar colecciÃ³n completa
    all_hadiths = []
    for hadiths in collections.values():
        all_hadiths.extend(hadiths)

    if all_hadiths:
        # Ordenar por ID
        all_hadiths.sort(key=lambda h: h.get("id", 0))

        # Guardar todos juntos
        all_path = OUTPUT_DIR / "hadiths_complete.json"
        with open(all_path, "w", encoding="utf-8") as f:
            json.dump(all_hadiths, f, ensure_ascii=False, indent=2)

        print(f"\nðŸ“¦ Guardado: {all_path} ({len(all_hadiths)} hadices)")

    # Guardar por colecciÃ³n
    for collection_key, hadiths in collections.items():
        # Ordenar por ID
        hadiths.sort(key=lambda h: h.get("id", 0))

        output_path = OUTPUT_DIR / f"{collection_key}.json"
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(hadiths, f, ensure_ascii=False, indent=2)

        print(f"  âœ“ {collection_key}.json ({len(hadiths)} hadices)")


def update_pubspec() -> None:
    """Actualiza pubspec.yaml para incluir los nuevos assets."""
    pubspec_path = Path("pubspec.yaml")

    if not pubspec_path.exists():
        print("\nâš ï¸  No se encontrÃ³ pubspec.yaml")
        return

    content = pubspec_path.read_text(encoding="utf-8")

    # Verificar si ya tiene la configuraciÃ³n de assets
    if "assets/hadiths/" in content:
        print("\nâœ… pubspec.yaml ya estÃ¡ configurado para hadiths")
        return

    # Buscar la secciÃ³n assets
    if "assets:" in content:
        # AÃ±adir assets/hadiths/ despuÃ©s de assets existentes
        new_content = content.replace(
            "assets:\n    - assets/audio/",
            "assets:\n    - assets/hadiths/\n    - assets/audio/"
        )

        if new_content == content:
            # Intentar otro patrÃ³n
            new_content = content.replace(
                "assets:",
                "assets:\n    - assets/hadiths/"
            )

        pubspec_path.write_text(new_content, encoding="utf-8")
        print("\nâœ… pubspec.yaml actualizado")
    else:
        print("\nâš ï¸  AÃ±ade manualmente a pubspec.yaml:")
        print("  assets:")
        print("    - assets/hadiths/")


def main():
    print("=" * 60)
    print("QiblaTime â€” Descargador de Hadices en EspaÃ±ol")
    print("=" * 60)
    print()

    # 1. Descargar Excel
    zip_data = download_excel()
    if not zip_data:
        print("\nâŒ No se pudo descargar el archivo")
        return

    # 2. Extraer Excel
    excel_data = extract_excel(zip_data)
    if not excel_data:
        print("\nâŒ No se pudo extraer el Excel")
        return

    # 3. Parsear Excel
    hadiths = parse_excel(excel_data)
    if not hadiths:
        print("\nâŒ No se encontraron hadices")
        return

    # 4. Agrupar por colecciÃ³n
    collections = group_by_collection(hadiths)
    print(f"\nðŸ“š Colecciones encontradas: {len(collections)}")

    # 5. Guardar JSON
    save_collections(collections)

    # 6. Actualizar pubspec.yaml
    update_pubspec()

    # Resumen
    print("\n" + "=" * 60)
    print("âœ… Â¡COMPLETADO!")
    print("=" * 60)
    print(f"\nðŸ“ Archivos generados en: {OUTPUT_DIR.absolute()}")
    print("\nðŸ“‹ Siguientes pasos:")
    print("   1. flutter clean")
    print("   2. flutter pub get")
    print("   3. flutter run")
    print("\nðŸ’¡ Para usar en tu app:")
    print("   - Carga desde assets/hadiths/{coleccion}.json")
    print("   - O usa assets/hadiths/hadiths_complete.json para todos")
    print()


if __name__ == "__main__":
    main()
